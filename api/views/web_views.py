from django.shortcuts import render, redirect
from api.models import *
from django.core.paginator import Paginator
from ..serializers import UsuarioComHistoricoSerializer
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import JsonResponse
from django.core.paginator import Paginator

def login(request):
    return render(request, 'index.html')


def home(request):
    users = CustomUser.objects.filter(is_adm=False).order_by("-pontuacao")[:5]

    userId = request.session.get('_auth_user_id')
    user = CustomUser.objects.filter(id=userId).first()
    primeiroAcesso = user.primeiroAcesso if user else False

    desafio_list = Desafio.objects.filter(idCampanha__isnull=True, is_active=True)

    desafio_paginator = Paginator(desafio_list, 5)
    desafio_page = request.GET.get('desafio_page')
    desafios = desafio_paginator.get_page(desafio_page)

    context = {
        'usuarios': users[1:],  
        'primeiro_usuario': users[0] if users else None,  
        'desafios': desafios,
        'primeiroAcesso': primeiroAcesso,
        'userId': userId    
    }

    return render(request, 'UserHtml/home.html', context)

def historicoCompra(request):
    eventos = Campanha.objects.filter(is_active=True)

    nome_query = request.GET.get('nome', '').strip()
    sort_by = request.GET.get('sort_by', 'dataCompra')
    order = request.GET.get('order', 'desc')

    # Filtra compras do usuário logado
    compra = Compra.objects.filter(idUsuario=request.user.id)

    # Filtro: somente por nome do produto (quando informado)
    if nome_query:
        compra_ids = ItensCompra.objects.filter(
            idProduto__nome__icontains=nome_query
        ).values_list('idCompra_id', flat=True)
        compra = compra.filter(id__in=compra_ids)

    # Ordenação
    if order == 'asc':
        compra = compra.order_by(sort_by)
    else:
        compra = compra.order_by(f'-{sort_by}')
        
    # Busca os itens das compras
    itensCompra = ItensCompra.objects.filter(idCompra__in=compra.values_list('id', flat=True))
    
    # Organiza os itens por compra
    compra_itens = {}
    for item in itensCompra:
        if item.idCompra_id not in compra_itens:
            compra_itens[item.idCompra_id] = {'itens': [], 'quantidadeItens': 0}
        compra_itens[item.idCompra_id]['itens'].append(item)
        compra_itens[item.idCompra_id]['quantidadeItens'] += item.qtdProduto

    # Paginação
    compra_paginator = Paginator(compra, 5)
    compra_page = request.GET.get('historicoCompra_page')
    compras = compra_paginator.get_page(compra_page)

    # Adiciona os itens e status ajustado
    for c in compras:
        c.itens = compra_itens.get(c.id, {}).get('itens', [])
        c.quantidadeItens = compra_itens.get(c.id, {}).get('quantidadeItens', 0)

        # Ajusta o status para "Em andamento" se atender à condição
        if c.entrega == 'Entrega' and c.pedido == 'Pendente' and c.obsEntrega != 'False':
            c.pedido = 'Em andamento'

    return render(request, 'UserHtml/historicoCompra.html', {
        'compra': compras,
        'eventos': eventos
    })

def primeiroAcesso(request):
    return render(request, 'primeiroAcesso.html')

def perfilUsuario(request):
    usuarioLogado = request.user
    serializer = UsuarioComHistoricoSerializer(usuarioLogado)
    dados_usuario = serializer.data

    context = {
        'historico': dados_usuario["ultimas_alteracoes"],  
        'saldoAtual': dados_usuario["ultimas_alteracoes"][0]
    }

    return render(request, 'UserHtml/perfilUsuario.html', context)

def listaProdutos(request):
    userId = request.session.get('_auth_user_id')
    user = CustomUser.objects.filter(id=userId).first()
    quantidade_moedas = user.saldo if user else 0

    search = request.GET.get("search", "")
    listaProdutos_list = Produto.objects.filter(is_active=True, idCampanha__isnull=True)
    if search:
        listaProdutos_list = listaProdutos_list.filter(nome__icontains=search)

    paginator = Paginator(listaProdutos_list, 8)
    page_number = request.GET.get("listaProduto_page")
    page_obj = paginator.get_page(page_number)

    # Construindo HTML dos produtos direto no Python
    html = ""
    for produto in page_obj:
        img_tag = f"<img class='imagensP-listaProdutos' src='{produto.img1.url}' alt='{produto.nome}'>" if produto.img1 else ""
        html += f"""
        <div class='imgD-listaProdutos' data-nome='{produto.nome.lower()}'>
            <button id='imgbotao' data-valor='{produto.id}'>{img_tag}</button>
        </div>
        """

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({"html": html})

    context = {
        "produtos": page_obj,
        "quantidade_moedas": quantidade_moedas,
        "search": search
    }
    return render(request, "UserHtml/listaProdutos.html", context)

def ProdutosListaCampanha(request, produto_id):
    userId = request.session.get('_auth_user_id')
    user = CustomUser.objects.filter(id=userId).first()
    quantidade_moedas = user.saldo if user else 0

    search = request.GET.get("search", "")
    listaProdutos_list = Produto.objects.filter(is_active=True, idCampanha__isnull=False, idCampanha=produto_id)
    if search:
        listaProdutos_list = listaProdutos_list.filter(nome__icontains=search)

    paginator = Paginator(listaProdutos_list, 8)
    page_number = request.GET.get("listaProduto_page")
    page_obj = paginator.get_page(page_number)

    # Construindo HTML dos produtos direto no Python
    html = ""
    for produto in page_obj:
        img_tag = f"<img class='imagensP-listaProdutos' src='{produto.img1.url}' alt='{produto.nome}'>" if produto.img1 else ""
        html += f"""
        <div class='imgD-listaProdutos' data-nome='{produto.nome.lower()}'>
            <button id='imgbotao' data-valor='{produto.id}'>{img_tag}</button>
        </div>
        """

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({"html": html})

    context = {
        "produtos": page_obj,
        "quantidade_moedas": quantidade_moedas,
        "search": search
    }
    return render(request, "UserHtml/produtosCampanha.html", context)

def ProdutosListaCampanhaAtivas(request):
    
    campanha = Campanha.objects.filter(is_active=True).order_by('nome')
    paginator = Paginator(campanha, 5)
    page_number = request.GET.get('campanha_page')
    page_obj = paginator.get_page(page_number)

    context = {
        "produtos": page_obj,
    }
    return render(request, "UserHtml/produtosCampanhasAtivas.html", context)



def cadastrarDesafio(request):
    campanhas = Campanha.objects.filter(is_active=True)
     
    return render(request, 'AdmHtml/cadastrarDesafio.html', {'campanhas': campanhas})

@login_required
def ranking(request):
    top_usuarios = CustomUser.objects.filter(is_adm=False).order_by('-pontuacao')[:7]
    
    usuario_logado = request.user
    
    usuario_em_top7 = any(usuario.id == usuario_logado.id for usuario in top_usuarios)
    
    todos_usuarios = CustomUser.objects.order_by('-saldo')
    posicao_usuario = 0
    for idx, usuario in enumerate(todos_usuarios, start=1):
        if usuario.id == usuario_logado.id:
            posicao_usuario = idx
            break
    
    context = {
        'top_usuarios': top_usuarios,
        'usuario_logado': {
            'id': usuario_logado.id,
            'first_name': usuario_logado.first_name,
            'saldo': usuario_logado.saldo,
            'posicao': posicao_usuario
        },
        'mostrar_usuario_logado': not usuario_em_top7 and posicao_usuario > 0
    }
    
    return render(request, 'UserHtml/ranking.html', context)

def listaEstoque(request):
    eventos = Campanha.objects.filter(is_active=True)
    
    # Get the search term from the URL (e.g., ?search=termo)
    search_term = request.GET.get('search', '')

    # Filter the products based on the search term and order them
    estoque_list = Produto.objects.filter(is_active=True, nome__icontains=search_term).order_by("id", "nome")

    # The rest of the pagination logic remains the same
    estoque_paginator = Paginator(estoque_list, 5) 
    estoque_page = request.GET.get('estoque_page') 
    estoque = estoque_paginator.get_page(estoque_page)

    return render(request, 'AdmHtml/listaEstoque.html', {'estoque': estoque, 'eventos': eventos})


def listaDeDesafios(request):
    search = request.GET.get('search', '').strip()

    desafios_queryset = Desafio.objects.filter(is_active=True).order_by('-id')

    if search:
        desafios_queryset = desafios_queryset.filter(nome__icontains=search)

    # Paginação
    desafio_paginator = Paginator(desafios_queryset, 5)
    desafio_page = request.GET.get('desafio_page')
    desafios = desafio_paginator.get_page(desafio_page)

    campanhas = Campanha.objects.filter(is_active=True)

    return render(request, 'AdmHtml/listaDeDesafios.html', {
        'desafios': desafios,
        'campanhas': campanhas,
        'search': search,   # passa pro template
    })

def listaDeUsuarios(request):
    # 1. Pega o parâmetro de busca 'q' da URL, o mesmo usado no formulário do HTML.
    search_query = request.GET.get('q', '') 
    
    # 2. Começa a query com todos os usuários não-administradores
    user_list = CustomUser.objects.filter(is_adm=False).order_by("first_name")

    # 3. Se houver um termo de busca, filtra o queryset ANTES da paginação
    if search_query:
        user_list = user_list.filter(
            Q(first_name__icontains=search_query) | # Busca por nome
            Q(username__icontains=search_query) |   # Busca por email
            Q(ra__icontains=search_query)           # Busca por RA
        )

    # 4. Aplica a paginação no resultado (já filtrado ou completo)
    user_paginator = Paginator(user_list, 5) # Aumentei para 15 por página, ajuste se desejar
    user_page = request.GET.get('user_page')
    usuarios = user_paginator.get_page(user_page)
    
    # Não há mais a parte de 'XMLHttpRequest', pois o JS vai recarregar a página inteira
    
    # 5. Envia os dados para o template
    context = {
        'usuarios': usuarios,
        'search_query': search_query  # Passa o termo da busca de volta para o template
    }
    
    return render(request, 'AdmHtml/listaDeUsuarios.html', context)


def desafiosCampanha(request, campanha_id):

    desafio = Desafio.objects.filter(idCampanha=campanha_id, is_active=True)
    desafio_paginator = Paginator(desafio, 5)
    desafio_page = request.GET.get('desafio_page')
    desafios = desafio_paginator.get_page(desafio_page)

    return render(request, 'UserHtml/desafiosCampanha.html', {'desafios': desafios})

# meu codigo primeiro
def desafiosCampanhaAtivas(request):
    campanha = Campanha.objects.filter(is_active=True).order_by('nome')
    campanha_paginator = Paginator(campanha, 5)
    campanha_page = request.GET.get('campanha_page')
    campanhas = campanha_paginator.get_page(campanha_page)

    return render(request, 'UserHtml/desafiosCampanhaAtivas.html', {'campanha': campanhas, 'campanhas': campanhas})



def listaDePedidos(request):
    status_pedido = request.GET.get('status')
    search = request.GET.get('search', '').strip()

    status_map = {
        '1': 'Concluído',
        '2': 'Pendente'
    }

    status = status_map.get(status_pedido)

    # Filtra as compras conforme o status
    if status:
        compras_queryset = Compra.objects.filter(pedido=status).order_by('-id')
    else:
        compras_queryset = Compra.objects.all().order_by('-id')

    # Filtro por nome do usuário (caso tenha search)
    if search:
        compras_queryset = compras_queryset.filter(idUsuario__first_name__icontains=search)

    # Pagina apenas as compras já filtradas
    compra_paginator = Paginator(compras_queryset, 5)
    compra_page = request.GET.get('compra_page')
    compras = compra_paginator.get_page(compra_page)

    # Busca os itens relacionados às compras paginadas
    pedidos = ItensCompra.objects.select_related('idProduto', 'idCompra').filter(idCompra__in=compras)

    return render(request, 'AdmHtml/listaDePedidos.html', {
        'compras': compras,
        'pedidos': pedidos,
        'search': search,  # <-- passa o termo pro template
    })



def carrinho(request):
    return render(request, 'UserHtml/carrinhoCompra.html')


def relatorio(request):
    return render(request, 'AdmHtml/relatorio.html')

def campanhas(request):

    campanhas = Campanha.objects.all()

    return render(request, 'AdmHtml/campanhas.html', {'campanhas': campanhas})

def teste(request):
    return render(request, 'UserHtml/teste.html')

def cadastrarUsuario(request):
    return render(request, 'AdmHtml/cadastrarUsuario.html')

def editarUsuario(request, id):
    
    userId = CustomUser.objects.filter(id=id).first()
    return render(request, 'AdmHtml/editarUsuario.html', {'userId': userId})


def adicionarMoedas(request):
    return render(request, 'AdmHtml/adicionarMoedas.html')


def exportar_vendas_excel(request):
    """
    View para exportar relatório de vendas em formato Excel
    """
    # Criar um novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Vendas"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        'ID Compra', 'Data da Compra', 'Cliente', 'RA Cliente', 
        'Produto', 'Quantidade', 'Valor Unitário', 'Valor Total Item',
        'Total da Compra', 'Status', 'Tipo de Entrega'
    ]
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados de vendas (compras) do banco de dados
    compras = Compra.objects.all().order_by('-dataCompra')
    
    row = 2
    print(compras)
    for compra in compras:
        print(compra)
        # Buscar itens da compra
        itens_compra = ItensCompra.objects.filter(idCompra=compra)
        
        for item in itens_compra:
            ws.cell(row=row, column=1, value=compra.id)
            ws.cell(row=row, column=2, value=compra.dataCompra.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=3, value=f"{compra.idUsuario.first_name} {compra.idUsuario.last_name}")
            ws.cell(row=row, column=4, value=compra.idUsuario.ra)
            ws.cell(row=row, column=5, value=item.idProduto.nome)
            ws.cell(row=row, column=6, value=item.qtdProduto)
            ws.cell(row=row, column=7, value=item.idProduto.valor)
            ws.cell(row=row, column=8, value=item.qtdProduto * item.idProduto.valor)
            ws.cell(row=row, column=9, value=compra.total)
            ws.cell(row=row, column=10, value=compra.pedido)
            ws.cell(row=row, column=11, value=compra.entrega)
            row += 1
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 25, 15, 30, 12, 15, 18, 18, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo com data atual
    filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    
    return response



def exportar_produtos_mais_vendidos_excel(request):
    """
    View para exportar relatório de produtos mais vendidos em formato Excel
    """
    # Criar um novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Mais Vendidos"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        'Produto', 'Quantidade Total Vendida', 'Valor Unitário', 
        'Receita Total', 'Número de Vendas'
    ]
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados de produtos mais vendidos
    from django.db.models import Sum, Count
    
    produtos_vendidos = (ItensCompra.objects
                        .select_related('idProduto')
                        .values('idProduto__nome', 'idProduto__valor')
                        .annotate(
                            quantidade_total=Sum('qtdProduto'),
                            numero_vendas=Count('idCompra', distinct=True)
                        )
                        .order_by('-quantidade_total'))
    
    row = 2
    for produto in produtos_vendidos:
        receita_total = produto['quantidade_total'] * produto['idProduto__valor']
        
        ws.cell(row=row, column=1, value=produto['idProduto__nome'])
        ws.cell(row=row, column=2, value=produto['quantidade_total'])
        ws.cell(row=row, column=3, value=produto['idProduto__valor'])
        ws.cell(row=row, column=4, value=receita_total)
        ws.cell(row=row, column=5, value=produto['numero_vendas'])
        row += 1
    
    # Ajustar largura das colunas
    column_widths = [30, 20, 15, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo com data atual
    filename = f"relatorio_produtos_mais_vendidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    
    return response


def exportar_usuarios_com_mais_moedas_excel(request):
    """
    View para exportar relatório de usuários com mais moedas em formato Excel
    """
    # Criar um novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários com Mais Moedas"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        'Posição', 'Nome Completo', 'RA', 'Email', 'Quantidade de Moedas'
    ]
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Buscar dados de usuários ordenados por quantidade de moedas (saldo)
    usuarios = CustomUser.objects.filter(is_adm=False).order_by('-saldo')
    
    row = 2
    posicao = 1
    for usuario in usuarios:
        ws.cell(row=row, column=1, value=posicao)
        ws.cell(row=row, column=2, value=f"{usuario.first_name} {usuario.last_name}")
        ws.cell(row=row, column=3, value=usuario.ra)
        ws.cell(row=row, column=4, value=usuario.email)
        ws.cell(row=row, column=5, value=usuario.saldo)
        row += 1
        posicao += 1
    
    # Ajustar largura das colunas
    column_widths = [10, 25, 15, 30, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo com data atual
    filename = f"relatorio_usuarios_mais_moedas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    
    return response

def desenvolvedores(request):
    devs = Desenvolvedores.objects.filter(professor=False)
    professores = Desenvolvedores.objects.filter(professor=True)
    return render(request, 'desenvolvedores.html', {'devs': devs, 'professores': professores})


